#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.sync_trades_from_binance import (
    BinanceTradeSync,
    build_monthly_export_output_path,
    build_summary_rows,
    normalize_trade_rows,
    write_trade_export_excel,
)


def sample_trades():
    return [
        {
            'time': 1711929600000,
            'symbol': 'BTCUSD_PERP',
            'side': 'BUY',
            'buyer': True,
            'maker': False,
            'price': '65000',
            'qty': '1',
            'quoteQty': '65000',
            'commission': '0.0001',
            'commissionAsset': 'BTC',
            'realizedPnl': '0.0010',
            'orderId': 1001,
            'id': 9001,
            'positionSide': 'LONG',
        },
        {
            'time': 1711933200000,
            'symbol': 'BTCUSD_PERP',
            'side': 'SELL',
            'buyer': False,
            'maker': True,
            'price': '65200',
            'qty': '2',
            'quoteQty': '130400',
            'commission': '0.0002',
            'commissionAsset': 'BTC',
            'realizedPnl': '-0.0005',
            'orderId': 1002,
            'id': 9002,
            'positionSide': 'SHORT',
        },
        {
            'time': 1711936800000,
            'symbol': 'BTCUSD_PERP',
            'side': 'SELL',
            'buyer': False,
            'maker': False,
            'price': '65300',
            'qty': '3',
            'quoteQty': '195900',
            'commission': '0.0003',
            'commissionAsset': 'BTC',
            'realizedPnl': '0.0003',
            'orderId': 1003,
            'id': 9003,
            'positionSide': 'SHORT',
        },
    ]


def test_normalize_trade_rows_returns_expected_columns():
    rows = normalize_trade_rows(sample_trades())

    assert len(rows) == 3
    assert rows[0]['trade_id'] == 9001
    assert rows[0]['trade_time'] == '2024-04-01 00:00:00'
    assert rows[0]['symbol'] == 'BTCUSD_PERP'
    assert rows[0]['price'] == 65000.0
    assert rows[1]['buyer'] is False
    assert rows[2]['realizedPnl'] == 0.0003


def test_build_summary_rows_groups_by_symbol_and_side():
    summary_rows = build_summary_rows(normalize_trade_rows(sample_trades()))

    assert len(summary_rows) == 2

    long_row = next(row for row in summary_rows if row['side'] == 'BUY')
    short_row = next(row for row in summary_rows if row['side'] == 'SELL')

    assert long_row['trade_count'] == 1
    assert long_row['total_qty'] == 1.0
    assert long_row['total_realized_pnl'] == 0.001

    assert short_row['trade_count'] == 2
    assert short_row['total_qty'] == 5.0
    assert short_row['total_quote_qty'] == 326300.0
    assert short_row['total_commission'] == 0.0005
    assert short_row['first_trade_time'] == '2024-04-01 01:00:00'
    assert short_row['last_trade_time'] == '2024-04-01 02:00:00'


def test_build_monthly_export_output_path_uses_exports_directory():
    start_label = '20260301'
    end_label = '20260331'

    output = build_monthly_export_output_path(
        base_dir=Path('/tmp/monipan1'),
        start_label=start_label,
        end_label=end_label,
    )

    assert output == Path('/tmp/monipan1/data/exports/binance_trades_20260301_20260331.xlsx')


def test_write_trade_export_excel_creates_expected_workbook(tmp_path):
    raw_rows = normalize_trade_rows(sample_trades())
    summary_rows = build_summary_rows(raw_rows)
    output_path = tmp_path / 'monthly_export.xlsx'

    write_trade_export_excel(output_path, raw_rows, summary_rows)

    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ['raw_trades', 'summary']
    assert workbook['raw_trades']['A1'].value == 'trade_time'
    assert workbook['summary']['A1'].value == 'symbol'
    assert workbook['summary']['C2'].value == 1


def test_export_monthly_trades_to_excel_uses_last_30_days_window(tmp_path):
    sync = BinanceTradeSync.__new__(BinanceTradeSync)
    sync.symbol = 'BTCUSD_PERP'
    sync.exchange = type('FakeExchange', (), {'is_connected': lambda self: True})()

    captured = {}

    def fake_fetch(start_time, end_time, raise_on_error=False):
        captured['start_time'] = start_time
        captured['end_time'] = end_time
        captured['raise_on_error'] = raise_on_error
        return sample_trades()

    sync._fetch_trades_batch = fake_fetch

    now = datetime(2026, 4, 1, 12, 0, 0)
    result = sync.export_monthly_trades_to_excel(base_dir=tmp_path, now=now)

    assert result['success'] is True
    assert result['trade_count'] == 3
    assert result['output_path'] == str(
        tmp_path / 'data/exports/binance_trades_20260302_20260401.xlsx'
    )
    assert Path(result['output_path']).exists()
    assert captured['end_time'] == now
    assert captured['start_time'] == now - timedelta(days=30)
    assert captured['raise_on_error'] is True


def test_export_monthly_trades_to_excel_requires_live_credentials(tmp_path):
    sync = BinanceTradeSync.__new__(BinanceTradeSync)
    sync.symbol = 'BTCUSD_PERP'
    sync.exchange = type(
        'FakeExchange',
        (),
        {
            'api_key': '',
            'api_secret': '',
            'is_connected': lambda self: False,
            'connect': lambda self: True,
        },
    )()

    result = sync.export_monthly_trades_to_excel(base_dir=tmp_path, now=datetime(2026, 4, 1, 12, 0, 0))

    assert result['success'] is False
    assert result['error'] == '缺少币安正式盘 API Key/Secret 配置'


def test_export_monthly_trades_to_excel_returns_error_when_fetch_fails(tmp_path):
    sync = BinanceTradeSync.__new__(BinanceTradeSync)
    sync.symbol = 'BTCUSD_PERP'
    sync.exchange = type(
        'FakeExchange',
        (),
        {
            'api_key': 'key',
            'api_secret': 'secret',
            'is_connected': lambda self: True,
        },
    )()

    def fake_fetch(start_time, end_time, raise_on_error=False):
        raise RuntimeError('API权限不足')

    sync._fetch_trades_batch = fake_fetch

    result = sync.export_monthly_trades_to_excel(base_dir=tmp_path, now=datetime(2026, 4, 1, 12, 0, 0))

    assert result['success'] is False
    assert result['error'] == 'API权限不足'
